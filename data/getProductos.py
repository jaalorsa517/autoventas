# -*- coding: utf-8 -*-

from openpyxl import load_workbook
from sqlite3 import Error

import sqlite3


class getProductos:
    def __init__(self, id=None, nom=None, precio=None, iva=None):
        self.id = id
        self.name = nom
        self.precio = precio
        self.iva = iva


def load():

    #CONSTANTES
    _PATH_XLXS = '/home/jaalorsa/Documentos/Drive/Gralac/Precios.xlsx'
    _PATH_DB = '/home/jaalorsa/Proyectos/flutter/autoventas/data/autoventasDB.db'
    _NUM_SHEET = 0
    _MAX_ROW = 277
    _MIN_ROW = 3

    #VARIABLES
    _id = 1001
    _list_producto = []
    _workbook = load_workbook(_PATH_XLXS)
    _sheet = _workbook.worksheets[_NUM_SHEET]

    print('INICIO PROCESO DE PRODUCTOS')
    print('OBTENIENDO DATOS DEL ARCHIVO XLSX')
    n = 0

    for colu in _sheet.iter_cols(
            min_col=1,
            max_col=9,
            min_row=_MIN_ROW,
            max_row=_MAX_ROW,
    ):
        for cell in colu:

            if (cell.column == 1 or cell.column == 2 or cell.column == 4
                    or cell.column == 9):

                if (cell.column == 1):
                    _list_producto.append(getProductos())
                    if (cell.value is not None):
                        _list_producto[cell.row - _MIN_ROW].id = int(
                            cell.value)
                    else:
                        _list_producto[cell.row - _MIN_ROW].id = 0

                if (cell.column == 2):
                    _list_producto[cell.row - _MIN_ROW].name = cell.value

                if (cell.column == 4):
                    _list_producto[cell.row - _MIN_ROW].iva = cell.value

                if (cell.column == 9):
                    _list_producto[cell.row - _MIN_ROW].precio = cell.value

                n += 1
                print(round(n * 100 / ((_MAX_ROW - 2) * _MIN_ROW), 1),
                      '%',
                      flush=True,
                      sep='',
                      end='\r')

    print()
    print('SE OBTUVO LOS DATOS DEL ARCHIVO XLSX')

    try:
        print('INICIO DE COPIADO A LA TABLA producto')
        conexion = sqlite3.connect(_PATH_DB)
        cursor = conexion.cursor()

        cursor.execute("DELETE FROM producto")
        conexion.commit()

        n = 0
        for prod in _list_producto:
            cursor.execute(
                "INSERT INTO producto (_id,id,nombre,valor,iva) VALUES ({},{},'{}',{},{})"
                .format(_id + n, prod.id, prod.name, prod.precio,
                        prod.iva * 100))
            conexion.commit()
            n += 1
            print(round(n * 100 / len(_list_producto), 1),
                  '%',
                  sep='',
                  flush=True,
                  end='\r')

    except Error:
        print()
        print(type(Error).__name__)

    finally:
        conexion.close()
        print()
        print('FIN DE PROCESO DE PRODUCTOS')


if __name__ == "__main__":
    load()
